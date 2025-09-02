import pandas as pd

def compute_correct(pred_col, gt_col):
    correct = []
    for p, g in zip(pred_col, gt_col):
        # caso X == X
        if p == g and pd.notna(p):
            correct.append(1)
        # caso entrambi vuoti o NaN
        elif (p in [None, ''] or pd.isna(p)) and (g in [None, ''] or pd.isna(g)):
            correct.append(1)
        else:
            correct.append(0)
    return correct


labels = ['INTENS','MESANGIALE','GRAN_GROSS','GRAN_FINE','GLOBAL_SEGMENTAL','PAR_IRREG','PAR_REGOL_CONT']

excel_path = "/work/grana_far2023_fomo/Pollastri_Glomeruli/Train_scripts/Results_folds_test/Prettified_scores_total_wsi_classification_Lv0_magistroni_norm_IF.xlsx"
output_path = "/work/grana_far2023_fomo/Pollastri_Glomeruli/Train_scripts/Results_folds_test/Prettified_scores_total_wsi_classification_Lv0_magistroni_norm_IF_out.xlsx"

df_excel = pd.read_excel(excel_path, skiprows=1)
#df_excel = pd.read_excel(excel_path)

df_excel["wsi_id_norm"] = (
    df_excel["wsi_id"]
    .str.replace("-", "_", regex=False)  
    .str.replace(" ", "_", regex=False) 
    .str.strip()
)

for label in labels:
    csv_path = f"/work/grana_far2023_fomo/Pollastri_Glomeruli/Train_scripts/Results_folds_test/{label}/allfolds_aggregated.csv"
    df_csv = pd.read_csv(csv_path)

 
    df_csv["G_ID_norm"] = (
        df_csv["G_ID"]
        .str.replace("-", "_", regex=False)  
        .str.replace(" ", "_", regex=False) 
        .str.replace("_FITC", "", regex=False) 
        .str.replace("_ind", "", regex=False)
        .str.replace("_IND", "", regex=False)
        .str.strip()
    )

  
    df_merged = df_excel.merge(
        df_csv, left_on="wsi_id_norm", right_on="G_ID_norm", how="left"
    )

    if label == 'INTENS':
        df_excel["intensity_preds"] = df_merged["Prediction_['INTENS']"]
    elif label == 'MESANGIALE': 
        df_excel["mesangial_preds"] = df_merged["Final_pred"].apply(lambda x: "X" if x == 1 else "")
        df_excel["mesangial_correct"] = compute_correct(df_excel["mesangial_preds"], df_excel["mesangial_gt"])
    elif label == 'GRAN_GROSS':
        df_excel["coarse_preds"] = df_merged["Final_pred"].apply(lambda x: "X" if x == 1 else "")
        df_excel["coarse_correct"] = compute_correct(df_excel["coarse_preds"], df_excel["coarse_gt"])
    elif label == 'GRAN_FINE':
        df_excel["fine_preds"] = df_merged["Final_pred"].apply(lambda x: "X" if x == 1 else "")
        df_excel["fine_correct"] = compute_correct(df_excel["fine_preds"], df_excel["fine_gt"])
    elif label == 'GLOBAL_SEGMENTAL':
        df_excel["segmental_preds"] = df_merged["Final_pred"].apply(lambda x: "X" if x == 'SEGM' else "")
        df_excel["segmental_correct"] = compute_correct(df_excel["segmental_preds"], df_excel["segmental_gt"])
        df_excel["global_preds"] = df_merged["Final_pred"].apply(lambda x: "X" if x == 'GLOB' else "")
        df_excel["global_correct"] = compute_correct(df_excel["global_preds"], df_excel["global_gt"])
    elif label == 'PAR_IRREG':
        df_excel["irregular_preds"] = df_merged["Final_pred"].apply(lambda x: "X" if x == 1 else "")
        df_excel["irregular_correct"] = compute_correct(df_excel["irregular_preds"], df_excel["irregular_gt"])
    elif label == 'PAR_REGOL_CONT':
        df_excel["continuous_preds"] = df_merged["Final_pred"].apply(lambda x: "X" if x == 1 else "")
        df_excel["continuous_correct"] = compute_correct(df_excel["continuous_preds"], df_excel["continuous_gt"])


df_excel = df_excel.drop(columns=["wsi_id_norm"], errors="ignore")

#df_excel.to_excel(output_path, index=False)
print(f"File aggiornato salvato in: {output_path}")

# Colora_celle_corrette
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook(output_path)
ws = wb.active

green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")


header = [cell.value for cell in ws[1]]
correct_cols = [
    idx+1 for idx, col in enumerate(header)
    if col.endswith("_correct") and col != "intensity_correct"
]

for row in ws.iter_rows(min_row=2):
    for idx in correct_cols:
        cell = row[idx-1]  
        if cell.value == 1:
            cell.fill = green_fill

#wb.save(output_path)
print(f"File finale con colori salvato in: {output_path}")


# CALCOLO METRICHE PER TUTTE LE LABEL

from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
import warnings

warnings.filterwarnings("ignore", category=FutureWarning)


labels_metrics = ["mesangial", "coarse", "fine", "segmental", "global", "irregular", "continuous"]

# funzione per trasformare X / vuoto in binario
def binarize(col):
    return col.apply(lambda x: 1 if x == "X" else 0)

print("\n--- METRICHE PER LABEL ---\n")
for label in labels_metrics:
    preds_col = f"{label}_preds"
    gt_col = f"{label}_gt"
    
    # binarizzo predizioni e ground truth
    df_excel[f"{label}_preds_bin"] = binarize(df_excel[preds_col])
    df_excel[f"{label}_gt_bin"] = binarize(df_excel[gt_col])
    
    y_true = df_excel[f"{label}_gt_bin"].astype(int)
    y_pred = df_excel[f"{label}_preds_bin"].astype(int)
    
    acc = accuracy_score(y_true, y_pred)
    prec = precision_score(y_true, y_pred, zero_division=0)
    rec = recall_score(y_true, y_pred, zero_division=0)
    f1 = f1_score(y_true, y_pred, zero_division=0)
    
    print(f"{label}: Accuracy={acc:.3f}, Precision={prec:.3f}, Recall={rec:.3f}, F1={f1:.3f}")
